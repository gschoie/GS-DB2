# -*- coding: utf-8 -*-
"""ECOS 주요국 환율(731Y001) → BOK_exchange_rates.xlsx 생성.

기존 Google Apps Script(ECOS 환율 프로젝트)를 이식한 것.
ECOS가 Google 데이터센터 IP를 차단해 UrlFetchApp이 "Address unavailable"로
실패하게 되면서 GitHub Actions + Python으로 옮겼다.

시트 구성: Daily / Monthly / Quarterly / Annual / Pivot Wide
- Monthly·Quarterly·Annual: 통화별 (평균, 기말)
- Pivot Wide: 행=통화×(평균|기말), 열=분기(1Q05…)와 연도(2005…) 혼합
"""

import datetime as dt
import os
import sys
import time
from zoneinfo import ZoneInfo

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ECOS_API_KEY = os.environ.get("ECOS_API_KEY", "")
START_DATE = "20050101"
STAT_CODE = "731Y001"

CURRENCIES = [
    {"code": "0000001", "key": "USD_KRW", "label": "원/달러"},
    {"code": "0000002", "key": "JPY100_KRW", "label": "원/엔"},
    {"code": "0000003", "key": "EUR_KRW", "label": "원/유로"},
    {"code": "0000053", "key": "CNY_KRW", "label": "원/중국위안"},
]

HEADER_FILL = PatternFill("solid", fgColor="CFE8F3")
NUM_FMT = "#,##0.0"
THIN_BORDER = Border(*(Side(style="thin"),) * 4)


def fetch_ecos_chunk(session, cycle, start, end, item_code):
    url = "/".join([
        "https://ecos.bok.or.kr/api/StatisticSearch",
        ECOS_API_KEY, "json", "kr", "1", "10000",
        STAT_CODE, cycle, start, end, item_code,
    ])
    max_attempts = 4
    last_error = None
    for attempt in range(1, max_attempts + 1):
        try:
            resp = session.get(url, timeout=30)
            if resp.status_code != 200:
                raise RuntimeError(f"HTTP {resp.status_code}: {resp.text[:500]}")
            data = resp.json()
            if isinstance(data.get("StatisticSearch", {}).get("row"), list):
                return data["StatisticSearch"]["row"]
            if data.get("RESULT", {}).get("CODE") == "INFO-200":  # 자료 없음
                return []
            raise RuntimeError(resp.text[:1000])
        except Exception as e:  # noqa: BLE001 - 재시도 후 최종 보고
            last_error = e
            print(f"ECOS 재시도 {attempt}/{max_attempts}: {item_code} / {start}~{end} / {e}")
            if attempt < max_attempts:
                time.sleep(attempt * 2)
    raise RuntimeError(f"ECOS 조회 최종 실패: {cycle} / {item_code} / {start}~{end}\n{last_error}")


def fetch_ecos(session, cycle, start, end, item_code):
    rows = []
    for year in range(int(start[:4]), int(end[:4]) + 1):
        chunk_start = start if year == int(start[:4]) else f"{year}0101"
        chunk_end = end if year == int(end[:4]) else f"{year}1231"
        rows.extend(fetch_ecos_chunk(session, cycle, chunk_start, chunk_end, item_code))
        time.sleep(0.25)
    return rows


def fetch_wide_map(cycle, start, end):
    """{period: {currency_key: value}}"""
    period_map = {}
    with requests.Session() as session:
        for currency in CURRENCIES:
            rows = fetch_ecos(session, cycle, start, end, currency["code"])
            print(f"{currency['label']}: {len(rows)}건")
            for row in rows:
                try:
                    value = float(str(row["DATA_VALUE"]).replace(",", ""))
                except (TypeError, ValueError):
                    continue
                if not row.get("TIME"):
                    continue
                period_map.setdefault(row["TIME"], {})[currency["key"]] = value
    return period_map


def get_quarter(yyyymmdd):
    return f"{(int(yyyymmdd[4:6]) + 2) // 3}Q{yyyymmdd[2:4]}"


def build_period_stats(daily_map, get_period):
    """{period: {key: {'avg': .., 'end': .., 'endDate': ..}}}"""
    buckets = {}
    for date in sorted(daily_map):
        period = get_period(date)
        bucket = buckets.setdefault(
            period,
            {c["key"]: {"sum": 0.0, "count": 0, "endDate": "", "end": None} for c in CURRENCIES},
        )
        for currency in CURRENCIES:
            value = daily_map[date].get(currency["key"])
            if value is None:
                continue
            b = bucket[currency["key"]]
            b["sum"] += value
            b["count"] += 1
            if not b["endDate"] or date > b["endDate"]:
                b["endDate"] = date
                b["end"] = value
    return {
        period: {
            key: {"avg": (b["sum"] / b["count"]) if b["count"] else None, "end": b["end"]}
            for key, b in bucket.items()
        }
        for period, bucket in buckets.items()
    }


def pivot_period_order(period):
    if "Q" in period:
        return int(f"20{period[2:4]}") * 10 + int(period[0])
    return int(period) * 10 + 5


def style_data_sheet(ws):
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            cell.number_format = NUM_FMT
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14


def write_wide_sheet(ws, daily_map):
    ws.append(["Date"] + [c["label"] for c in CURRENCIES])
    for date in sorted(daily_map):
        ws.append([f"{date[:4]}-{date[4:6]}-{date[6:8]}"]
                  + [daily_map[date].get(c["key"]) for c in CURRENCIES])
    style_data_sheet(ws)


def write_period_stats_sheet(ws, stats, first_header, sort_key=None):
    ws.append([first_header] + [f"{c['label']}({kind})" for c in CURRENCIES for kind in ("평균", "기말")])
    for period in sorted(stats, key=sort_key):
        label = f"{period[:4]}-{period[4:6]}" if first_header == "Month" else period
        ws.append([label] + [stats[period][c["key"]][kind] for c in CURRENCIES for kind in ("avg", "end")])
    style_data_sheet(ws)


def write_pivot_sheet(ws, daily_map):
    quarterly = build_period_stats(daily_map, get_quarter)
    annual = build_period_stats(daily_map, lambda d: d[:4])
    stats = {**quarterly, **annual}
    periods = sorted(stats, key=pivot_period_order)

    ws.append(["데이터"] + periods)
    for currency in CURRENCIES:
        ws.append([f"{currency['label']}(평균)"] + [stats[p][currency["key"]]["avg"] for p in periods])
        ws.append([f"{currency['label']}(기말)"] + [stats[p][currency["key"]]["end"] for p in periods])

    ws.freeze_panes = "B2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        is_end_row = row_idx % 2 == 1  # 3, 5, 7… = 기말 행
        row[0].font = Font(bold=True)
        for cell in row[1:]:
            cell.number_format = NUM_FMT
            cell.alignment = Alignment(horizontal="right")
            if is_end_row:
                cell.font = Font(color="FF0000")
    ws.column_dimensions["A"].width = 18
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 11


def main():
    if not ECOS_API_KEY:
        sys.exit("ECOS_API_KEY 환경변수가 없습니다.")
    end_date = dt.datetime.now(ZoneInfo("Asia/Seoul")).strftime("%Y%m%d")
    print(f"조회 구간: {START_DATE} ~ {end_date}")

    daily_map = fetch_wide_map("D", START_DATE, end_date)
    if not daily_map:
        sys.exit("ECOS에서 내려받은 일별 환율 자료가 없습니다.")

    monthly = build_period_stats(daily_map, lambda d: d[:6])
    quarterly = build_period_stats(daily_map, get_quarter)
    annual = build_period_stats(daily_map, lambda d: d[:4])

    wb = Workbook()
    write_wide_sheet(wb.active, daily_map)
    wb.active.title = "Daily"
    write_period_stats_sheet(wb.create_sheet("Monthly"), monthly, "Month")
    write_period_stats_sheet(wb.create_sheet("Quarterly"), quarterly, "Quarter", sort_key=pivot_period_order)
    write_period_stats_sheet(wb.create_sheet("Annual"), annual, "Year")
    write_pivot_sheet(wb.create_sheet("Pivot Wide"), daily_map)

    out_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "..", "telegram_research_dashboard", "static", "BOK_exchange_rates.xlsx",
    )
    out_path = os.path.normpath(out_path)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f"엑셀 파일 생성 완료: {out_path} ({len(daily_map)}일)")


if __name__ == "__main__":
    main()
