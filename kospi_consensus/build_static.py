# -*- coding: utf-8 -*-
"""SQLite 스냅샷 → 'consensus_revision.html' + 'consensus_full.xlsx' 생성.

- 화면: (리비전 주) 어닝 변화 큰 순 상위 20 — 컨센 변동 + 주가 변동 함께 · (첫 주) 컨센 레벨 상위 20
- 전체 종목 데이터는 엑셀 다운로드 버튼으로 제공
출력: telegram_research_dashboard/static/{consensus_revision.html, consensus_full.xlsx}
"""
import datetime as dt
import json
import os
import sys
from collections import Counter

import openpyxl
from openpyxl.styles import Font, PatternFill

import db

sys.stdout.reconfigure(encoding="utf-8")

BASE = os.path.dirname(os.path.abspath(__file__))
STATIC = os.path.normpath(os.path.join(BASE, "..", "telegram_research_dashboard", "static"))
OUT_HTML = os.path.join(STATIC, "consensus_revision.html")
OUT_XLSX = os.path.join(STATIC, "consensus_full.xlsx")
SECTORS_PATH = os.path.join(BASE, "sectors.json")   # 코드→섹터(세부업종) 매핑
MAX_PERIODS = 10                                     # 화면 주차 이동 최대 비교주 수
MONTHS3 = 3                                           # 장기 비교 기준(개월)

HORIZONS = [
    ("이번분기", "quarter", None, "q"),
    ("2026E", "annual", "2026.12", "a26"),
    ("2027E", "annual", "2027.12", "a27"),
    ("2028E", "annual", "2028.12", "a28"),
]
PAGE_KEYS = ["q", "a26", "a27"]


def dominant_quarter(con, snap):
    rows = con.execute(
        "SELECT period FROM consensus_snapshots WHERE snapshot_date=? AND kind='quarter'",
        (snap,)).fetchall()
    return Counter(r[0] for r in rows).most_common(1)[0][0] if rows else None


def series(con, snap, base, kind, period):
    cur = {r["code"]: (r["name"], r["op_profit"]) for r in con.execute(
        "SELECT code,name,op_profit FROM consensus_snapshots "
        "WHERE snapshot_date=? AND kind=? AND period=? AND op_profit IS NOT NULL",
        (snap, kind, period))}
    prev = {}
    if base:
        prev = {r["code"]: r["op_profit"] for r in con.execute(
            "SELECT code,op_profit FROM consensus_snapshots "
            "WHERE snapshot_date=? AND kind=? AND period=? AND op_profit IS NOT NULL",
            (base, kind, period))}
    out = {}
    for code, (name, curr) in cur.items():
        b = prev.get(code)
        wow = ((curr - b) / abs(b) * 100.0) if (b not in (None, 0)) else None
        out[code] = {"code": code, "name": name, "curr": curr, "base": b, "wow": wow}
    return out


def price_date_of(con, date):
    r = con.execute(
        "SELECT price_date, COUNT(*) c FROM stock_prices WHERE snapshot_date=? "
        "GROUP BY price_date ORDER BY c DESC LIMIT 1", (date,)).fetchone()
    return r["price_date"] if r else None


def report_week(date_str):
    """주간 버킷을 '화~월'로 잡아 그 주의 시작 화요일(정렬 가능한 키)을 반환.

    정기 실행은 금요일이지만, 주말에 안 돌고 월요일 아침에 지각 실행할 수 있다.
    화~월로 묶으면 그 금요일과 다음 월요일이 같은 주가 되어(둘 다 시작 화요일이 같음)
    지각분이 새 주가 아니라 '그 주 최신'으로 처리된다.
    """
    y, m, d = map(int, date_str.split("-"))
    dd = dt.date(y, m, d)
    start = dd - dt.timedelta(days=(dd.weekday() - 1) % 7)   # weekday: Mon=0..Sun=6, 직전(포함) 화요일
    return start.isoformat()


def weekly_reps(dates):
    """오름차순 스냅샷 날짜 → 주(화~월)별 대표(그 주 최신) 날짜 리스트(오름차순).

    같은 주에 여러 번 돌려도 마지막(가장 늦은) 실행분만 그 주 대표로 쓰인다.
    연속한 두 대표가 곧 '전주 → 이번주' 비교쌍이 된다.
    """
    latest_of_week = {}
    for d in dates:                 # dates 오름차순 → 같은 주는 뒤 날짜가 덮어써 대표가 됨
        latest_of_week[report_week(d)] = d
    return [latest_of_week[w] for w in sorted(latest_of_week)]


def load_sectors():
    try:
        with open(SECTORS_PATH, encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return {}


def load_universe_meta():
    """code → {mkt(시장), groups[...], cov(커버리지 여부)}. universe.json 기반."""
    try:
        with open(os.path.join(BASE, "data", "universe.json"), encoding="utf-8") as f:
            u = json.load(f)
    except FileNotFoundError:
        return {}
    out = {}
    for code, v in u.items():
        if isinstance(v, dict):
            g = v.get("groups", [])
            out[code] = {"mkt": v.get("market"), "groups": g, "cov": "커버리지" in g}
    return out


def three_month_ref(snap, reps, months=MONTHS3):
    """snap보다 앞선 rep 중 'snap - months*30일'에 가장 가까운 날짜.

    아직 months치 히스토리가 없으면(타깃이 모든 rep보다 이르면) 가장 오래된 rep이
    자동 선택된다(헤더에 실제 날짜를 표기하므로 오해 없음). 앞선 rep이 없으면 None.
    """
    earlier = [r for r in reps if r < snap]
    if not earlier:
        return None
    y, m, d = map(int, snap.split("-"))
    target = dt.date(y, m, d) - dt.timedelta(days=months * 30)
    return min(earlier, key=lambda r: abs((dt.date(*map(int, r.split("-"))) - target).days))


def _counts(vals):
    rated = [v for v in vals if v is not None]
    up = sum(1 for v in rated if v > 0.05)
    down = sum(1 for v in rated if v < -0.05)
    return up, down, len(rated) - up - down


def build_period(con, snap, base, ref3, sectors, umeta):
    """한 비교주: 전주(base)와 3개월전(ref3) 두 기준 대비 값을 함께 담는다."""
    qp = dominant_quarter(con, snap)
    psnap = db.price_map(con, snap)
    pbase = db.price_map(con, base) if base else {}
    pref3 = db.price_map(con, ref3) if ref3 else {}

    def pw(mp, code):
        a, b = psnap.get(code), mp.get(code)
        return ((a - b) / b * 100.0) if (a and b) else None

    hz_meta = {k: (kind, period) for _, kind, period, k in HORIZONS}
    labels = {k: l for l, _, _, k in HORIZONS}
    names = {}
    hz = []
    for key in PAGE_KEYS:
        kind, period = hz_meta[key]
        p = period or qp
        sw = series(con, snap, base, kind, p)                     # 전주 대비
        s3 = series(con, snap, ref3, kind, p) if ref3 else {}     # 3개월전 대비
        for code, d in sw.items():
            names[code] = d["name"]
        uw = _counts([d["wow"] for d in sw.values()])
        u3 = _counts([d["wow"] for d in s3.values()]) if ref3 else (0, 0, 0)
        rows = []
        for code, d in sw.items():
            d3 = s3.get(code)
            um = umeta.get(code, {})
            rows.append({"code": code, "name": d["name"], "sec": sectors.get(code, "기타"),
                         "mkt": um.get("mkt"), "cov": bool(um.get("cov")),
                         "grp": ",".join(um.get("groups", [])),
                         "curr": d["curr"], "base": d["base"], "wow": d["wow"], "pwow": pw(pbase, code),
                         "base3": d3["base"] if d3 else None, "wow3": d3["wow"] if d3 else None,
                         "pwow3": pw(pref3, code)})
        if base:
            rows.sort(key=lambda r: (r["wow"] is None, -abs(r["wow"]) if r["wow"] is not None else 0))
        else:
            rows.sort(key=lambda r: -(r["curr"] or 0))
        hz.append({"label": labels[key], "period": p, "key": key,
                   "up": uw[0], "down": uw[1], "flat": uw[2],
                   "up3": u3[0], "down3": u3[1], "flat3": u3[2], "rows": rows})
    uni = con.execute("SELECT COUNT(*) FROM universe WHERE snapshot_date=?", (snap,)).fetchone()[0]
    return {"snapshot_date": snap, "base_date": base, "ref3_date": ref3,
            "price_date_snap": price_date_of(con, snap),
            "price_date_base": price_date_of(con, base) if base else None,
            "price_date_ref3": price_date_of(con, ref3) if ref3 else None,
            "universe": uni, "covered": len(names), "horizons": hz}


def build():
    con = db.connect()
    dates = db.snapshot_dates(con)
    if not dates:
        print("스냅샷이 없습니다. run_snapshot.py 먼저 실행.")
        return
    sectors = load_sectors()
    umeta = load_universe_meta()
    reps = weekly_reps(dates)
    snap = reps[-1]
    base = reps[-2] if len(reps) >= 2 else None

    # --- 엑셀: 최신 주 기준 전체 종목(a28 포함) ---
    qp = dominant_quarter(con, snap)
    psnap = db.price_map(con, snap)
    pbase = db.price_map(con, base) if base else {}

    def pwow(code):
        a, b = psnap.get(code), pbase.get(code)
        return ((a - b) / b * 100.0) if (a and b) else None

    maps, periods = {}, {}
    for label, kind, period, key in HORIZONS:
        p = period or qp
        periods[key] = p
        maps[key] = series(con, snap, base, kind, p)
    names = {}
    for key in maps:
        for code, d in maps[key].items():
            names[code] = d["name"]
    stocks = []
    for code in sorted(names):
        um = umeta.get(code, {})
        row = {"code": code, "name": names[code],
               "mkt": um.get("mkt"), "sec": sectors.get(code, "기타"),
               "grp": ",".join(um.get("groups", [])),
               "pcur": psnap.get(code), "pbase": pbase.get(code), "pw": pwow(code)}
        for key in ("q", "a26", "a27", "a28"):
            d = maps[key].get(code)
            row[key] = d["curr"] if d else None
            row[key + "_b"] = d["base"] if d else None
            row[key + "_w"] = d["wow"] if d else None
        stocks.append(row)
    write_excel(stocks, snap, base, periods)

    # --- 화면: 최근 여러 주 비교(주차 이동 ◀▶) ---
    if len(reps) == 1:
        pairs = [(reps[0], None)]
    else:
        pairs = [(reps[i], reps[i - 1]) for i in range(len(reps) - 1, 0, -1)][:MAX_PERIODS]
    periods_payload = [build_period(con, s, b, three_month_ref(s, reps), sectors, umeta)
                       for s, b in pairs]
    sec_list = sorted({r["sec"] for pp in periods_payload for h in pp["horizons"] for r in h["rows"]})

    payload = {"has_revision": base is not None,
               "periods": periods_payload, "sectors": sec_list,
               "groups": ["KOSPI200", "KOSDAQ50", "커버리지"], "markets": ["코스피", "코스닥"]}
    os.makedirs(STATIC, exist_ok=True)
    with open(OUT_HTML, "w", encoding="utf-8") as f:
        gen = dt.datetime.now(dt.timezone(dt.timedelta(hours=9))).strftime("%Y-%m-%d %H:%M")
        f.write(_TEMPLATE
                .replace("__DATA__", json.dumps(payload, ensure_ascii=False).replace("</", "<\\/"))
                .replace("__GENERATED__", gen))
    con.close()
    print(f"생성 완료 [{len(periods_payload)}개 비교주 · {'리비전' if base else '베이스라인'}]: {OUT_HTML}\n            엑셀: {OUT_XLSX}")


def write_excel(stocks, snap, base, periods):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "컨센"
    labels = [("q", "이번분기"), ("a26", "2026E"), ("a27", "2027E"), ("a28", "2028E")]
    header = ["종목코드", "종목명", "시장", "섹터", "그룹"]
    cols = []
    for key, lab in labels:
        per = periods.get(key, "")
        header.append(f"{lab}({per}) 영업이익")
        cols.append(("v", key))
        if base:
            header += [f"{lab} 전주", f"{lab} 컨센변동%"]
            cols += [("b", key), ("w", key)]
    header.append("현재 종가")
    cols.append(("pc", None))
    if base:
        header += ["전주 종가", "주가변동%"]
        cols += [("pb", None), ("pw", None)]

    title = (f"영업이익 컨센(KOSPI200·KOSDAQ50·커버리지) · 기준일 {snap}"
             + (f" · 전주 {base}" if base else " · 첫 스냅샷"))
    ws.append([title])
    ws.append(header)
    for st in stocks:
        r = [st["code"], st["name"], st.get("mkt"), st.get("sec"), st.get("grp")]
        for typ, key in cols:
            if typ == "v":
                r.append(st[key])
            elif typ == "b":
                r.append(st[key + "_b"])
            elif typ == "w":
                w = st[key + "_w"]
                r.append(round(w, 1) if w is not None else None)
            elif typ == "pc":
                r.append(st["pcur"])
            elif typ == "pb":
                r.append(st["pbase"])
            else:
                r.append(round(st["pw"], 1) if st["pw"] is not None else None)
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=12)
    for c in ws[2]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="E8EEF5")
    ws.freeze_panes = "A3"
    widths = [10, 16, 8, 14, 22] + [15] * (len(header) - 5)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    wb.save(OUT_XLSX)


_TEMPLATE = r"""<!doctype html>
<html lang="ko"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>KOSPI200 컨센 리비전</title>
<style>
:root{--bg:#f6f7f9;--card:#fff;--ink:#1a1d21;--muted:#6b7280;--line:#e5e7eb;
--up:#0ca30c;--down:#d03b3b;--accent:#185fa5}
@media(prefers-color-scheme:dark){:root{--bg:#141516;--card:#1d1e1f;--ink:#f2f2f0;
--muted:#9a9a92;--line:#2f3032}}
*{box-sizing:border-box}body{margin:0;background:var(--bg);color:var(--ink);
font-family:-apple-system,"Segoe UI","Malgun Gothic",sans-serif;line-height:1.5}
.wrap{max-width:860px;margin:0 auto;padding:24px 18px 60px}
h1{font-size:20px;font-weight:600;margin:0 0 2px}.sub{color:var(--muted);font-size:13px}
.sched{display:inline-block;background:var(--card);border:1px solid var(--line);border-radius:6px;
padding:3px 9px;margin:6px 0 0;font-size:11px;color:var(--muted)}
.sched b{color:var(--ink);font-weight:600}
.schedq{opacity:.75}
@media(max-width:620px){.schedq{display:none}}
.status{display:flex;flex-wrap:wrap;gap:8px 16px;align-items:center;background:var(--card);
border:1px solid var(--line);border-radius:12px;padding:12px 16px;margin:14px 0 16px;font-size:13px}
.status b{font-weight:600}
.dl{display:inline-flex;align-items:center;gap:6px;background:var(--accent);color:#fff;
text-decoration:none;font-size:13px;font-weight:600;padding:9px 15px;border-radius:9px;margin-bottom:22px}
.cards{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:12px;margin-bottom:22px}
.card{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:13px 15px}
.card .lab{font-size:12px;color:var(--muted);margin-bottom:7px}
.card .nums{display:flex;gap:14px;align-items:baseline;font-size:21px;font-weight:600}
.up{color:var(--up)}.down{color:var(--down)}
.bar{display:flex;height:5px;border-radius:3px;overflow:hidden;margin:9px 0 4px;background:var(--line)}
.bar i{display:block}.bar .iu{background:var(--up)}.bar .id{background:var(--down)}
.tiny{font-size:11px;color:var(--muted);font-weight:400}
.tabs{display:flex;gap:6px;margin:6px 0 12px}
.tab{border:1px solid var(--line);background:var(--card);color:var(--ink);font-size:13px;
padding:7px 14px;border-radius:8px;cursor:pointer}
.tab.on{background:var(--accent);color:#fff;border-color:var(--accent);font-weight:600}
.tblwrap{overflow-x:auto}
table{width:100%;border-collapse:collapse;font-size:13px;background:var(--card);
border:1px solid var(--line);border-radius:12px;overflow:hidden;min-width:560px}
th,td{padding:7px 11px;text-align:right;border-bottom:1px solid var(--line);white-space:nowrap}
th:first-child,td:first-child{text-align:center;color:var(--muted);width:30px}
th.l,td.l{text-align:left}td.sec{color:var(--muted);font-size:12px}
th{color:var(--muted);font-weight:500;font-size:12px}tr:last-child td{border-bottom:0}
th.sortable{cursor:pointer;user-select:none}th.sortable:hover{color:var(--ink)}
a.stk{color:var(--ink);text-decoration:none;border-bottom:1px dotted var(--muted)}
a.stk:hover{color:var(--accent);border-color:var(--accent)}
.lim{margin:0 0 10px;border:1px solid var(--line);background:var(--card);color:var(--accent);
font-size:12px;font-weight:600;padding:6px 12px;border-radius:8px;cursor:pointer}
.wknav{display:flex;align-items:center;gap:10px;flex-wrap:wrap;margin:2px 0 16px}
.wkbtn{border:1px solid var(--line);background:var(--card);color:var(--ink);font-size:13px;
font-weight:600;padding:7px 13px;border-radius:8px;cursor:pointer}
.wkbtn:disabled{opacity:.38;cursor:default}.wklab{font-size:14px;font-weight:600}
.ctlbar{margin:2px 0 10px}.ctllab{font-size:12px;color:var(--muted)}
.secfil{font-size:13px;color:var(--ink);background:var(--card);border:1px solid var(--line);
border-radius:8px;padding:6px 10px;margin-left:6px}
.modetog{display:inline-flex;margin-left:14px;border:1px solid var(--line);border-radius:8px;
overflow:hidden;vertical-align:middle}
.modetog button{border:0;background:var(--card);color:var(--ink);font-size:12px;font-weight:600;
padding:6px 13px;cursor:pointer}.modetog button.on{background:var(--accent);color:#fff}
td.sec .secpick{cursor:pointer}td.sec .secpick:hover{color:var(--accent);text-decoration:underline}
.clrfil{margin-left:8px;border:1px solid var(--line);background:var(--card);color:var(--down);
font-size:12px;font-weight:600;padding:6px 11px;border-radius:8px;cursor:pointer}
.pos{color:var(--up);font-weight:600}.neg{color:var(--down);font-weight:600}
.note{background:var(--card);border:1px dashed var(--line);border-radius:12px;
padding:14px 16px;color:var(--muted);font-size:13px;margin-bottom:18px}
td.mktc{text-align:center}
.mkt-kp,.mkt-kq{font-size:10px;font-weight:600;padding:2px 6px;border-radius:6px;white-space:nowrap}
.mkt-kp{color:#185fa5;background:rgba(24,95,165,.12)}
.mkt-kq{color:#b0700a;background:rgba(176,112,10,.14)}
.covstar{color:#e0a400;font-size:11px}
.grpfil{font-size:13px;color:var(--ink);background:var(--card);border:1px solid var(--line);
border-radius:8px;padding:6px 10px;margin-left:6px}
</style></head><body><div class="wrap">
<h1>영업이익 컨센 리비전 <span class="tiny" style="font-weight:400">KOSPI200 · KOSDAQ50 · 리서치커버리지</span></h1>
<div class="sub" id="sub"></div>
<div class="sched">🕙 정기 업데이트 <b>매주 금요일 오후 5시</b>(한국시간) · 실제 갱신 <b>__GENERATED__</b><span class="schedq"> — 대기열 지연으로 예약보다 늦을 수 있음</span></div>
<div class="status" id="status"></div>
<a class="dl" href="consensus_full.xlsx" download>📥 전체 종목 엑셀 다운로드</a>
<div class="wknav" id="wknav"></div>
<div id="cards"></div>
<div class="tabs" id="tabs"></div>
<div class="ctlbar" id="ctlbar"></div>
<div class="tblwrap"><div id="tbl"></div></div>
<div class="tiny" id="foot" style="margin-top:14px"></div>
</div>
<script>
var D=__DATA__;
var pi=0, sel=0, sortKey=null, sortDir=-1, showAll=false, secFilter='', grpFilter='', MODE='wk', TOP=30;
function P(){return D.periods[pi]}
function $(id){return document.getElementById(id)}
function m3(){return MODE==='3m'}
function chgField(){return m3()?'wow3':'wow'}
function fmt(n){return n==null?'-':Math.round(n).toLocaleString('ko-KR')}
function pct(w){if(w==null)return'<span class="tiny">-</span>';
 var c=w>0?'pos':(w<0?'neg':'');return'<span class="'+c+'">'+(w>0?'+':'')+w.toFixed(1)+'%</span>'}
function mmdd(s){return s?s.slice(5).replace('-','/'):''}   // 'YYYY-MM-DD' → 'MM/DD'
function qq(p){if(!p)return'';var m=p.slice(5,7),y=p.slice(2,4);   // '2026.06' → '2Q26'
 return ({'03':'1Q','06':'2Q','09':'3Q','12':'4Q'}[m]||m)+y}
function plab(h){return h.key==='q'?qq(h.period):h.period}   // 분기는 'nQyy', 연간은 'YYYY.12'
function nameCell(d){return '<a class="stk" href="https://finance.naver.com/item/fchart.naver?code='+
 d.code+'" target="_blank" rel="noopener">'+d.name+'</a>'+(d.cov?' <span class="covstar" title="리서치 커버리지">★</span>':'')}
function mktBadge(m){return m?'<span class="'+(m==='코스닥'?'mkt-kq':'mkt-kp')+'">'+m+'</span>':''}
function sortRows(rows){
 if(sortKey==null){var kk=chgField();   // 기본: 선택 기준(주간/3개월) 변화 큰 순
  return rows.slice().sort(function(a,b){var x=a[kk],y=b[kk];
   x=x==null?-1:Math.abs(x);y=y==null?-1:Math.abs(y);return y-x});}
 return rows.slice().sort(function(a,b){var x=a[sortKey],y=b[sortKey];
  if(x==null&&y==null)return 0;if(x==null)return 1;if(y==null)return -1;   // 값 없는 종목은 뒤로
  return typeof x==='string'?sortDir*x.localeCompare(y,'ko'):sortDir*(x-y)});
}
function arrow(k){return sortKey===k?(sortDir<0?' ▼':' ▲'):''}

function renderNav(){   // 주차 이동 ◀▶ (비교주가 2개 이상일 때만)
 var n=D.periods.length, p=P(), el=$('wknav');
 if(n<2){el.innerHTML='';return}
 var older=pi<n-1, newer=pi>0;
 el.innerHTML='<button class="wkbtn" id="wkprev"'+(older?'':' disabled')+'>◀ 이전주</button>'+
  '<span class="wklab">전주 '+p.base_date+' → 현재 '+p.snapshot_date+'</span>'+
  '<button class="wkbtn" id="wknext"'+(newer?'':' disabled')+'>다음주 ▶</button>'+
  '<span class="tiny">'+(pi+1)+'/'+n+'</span>';
 if(older)$('wkprev').onclick=function(){pi++;render()};
 if(newer)$('wknext').onclick=function(){pi--;render()};
}
function renderTop(){
 var p=P();
 $('sub').textContent=D.has_revision?('전주 '+p.base_date+' → 현재 '+p.snapshot_date):(p.snapshot_date+' 기준 · 첫 스냅샷');
 $('status').innerHTML='<span>📅 기준일 <b>'+p.snapshot_date+'</b></span>'+
  '<span>🎯 유니버스 <b>'+p.universe+'</b></span>'+
  '<span>✅ 컨센 확보 <b>'+p.covered+'</b> · 없음 '+(p.universe-p.covered)+'</span>';
 if(D.has_revision){
  var M=m3(), ch='<div class="cards">';
  p.horizons.forEach(function(h){var up=M?h.up3:h.up,dn=M?h.down3:h.down,fl=M?h.flat3:h.flat,t=up+dn+fl||1;
   ch+='<div class="card"><div class="lab">'+h.label+' ('+plab(h)+')</div>'+
    '<div class="nums"><span class="up">'+up+' ▲</span><span class="down">'+dn+' ▼</span></div>'+
    '<div class="bar"><i class="iu" style="width:'+(up/t*100)+'%"></i>'+
    '<i class="id" style="width:'+(dn/t*100)+'%"></i></div>'+
    '<div class="tiny">보합 '+fl+' · 평가 '+(up+dn+fl)+'</div></div>';});
  $('cards').innerHTML=ch+'</div>';
  var bd=M?p.ref3_date:p.base_date, pbd=M?p.price_date_ref3:p.price_date_base;
  $('foot').innerHTML='컨센 변화 = '+(M?'3개월전':'전주')+'('+(bd||'')+') 대비 영업이익 컨센 변화율 · 주가 변동 = 종가 '+
   (pbd||'')+' → '+(p.price_date_snap||'')+' · 상위 30 기본(전체 보기·열 정렬) · 섹터/종목명 클릭 활용';
 }else{
  $('cards').innerHTML='<div class="note">첫 스냅샷이라 전주 대비 변동은 다음 스냅샷부터 표시됩니다. 아래는 현재 컨센 레벨이며, 전체 종목은 엑셀에서 확인하세요.</div>';
  $('foot').innerHTML='';
 }
}
function renderCtl(){   // 그룹/섹터 필터 + 기준(주간/3개월) 토글
 var gopts='<option value="">전체 그룹</option>';
 (D.groups||[]).forEach(function(g){gopts+='<option value="'+g+'"'+(g===grpFilter?' selected':'')+'>'+g+'</option>'});
 var opts='<option value="">전체 섹터</option>';
 D.sectors.forEach(function(s){opts+='<option value="'+s+'"'+(s===secFilter?' selected':'')+'>'+s+'</option>'});
 var clr=(secFilter||grpFilter)?'<button class="clrfil" id="clrfil">✕ 필터 해제</button>':'';
 var tog=D.has_revision?'<span class="modetog"><button data-m="wk"'+(MODE==='wk'?' class="on"':'')+'>주간</button>'+
  '<button data-m="3m"'+(MODE==='3m'?' class="on"':'')+'>3개월</button></span>':'';
 $('ctlbar').innerHTML='<label class="ctllab">그룹 <select class="grpfil" id="grpfil">'+gopts+'</select></label>'+
  '<label class="ctllab" style="margin-left:6px">섹터 <select class="secfil" id="secfil">'+opts+'</select></label>'+clr+tog;
 $('grpfil').onchange=function(){grpFilter=this.value;renderCtl();draw()};
 $('secfil').onchange=function(){secFilter=this.value;renderCtl();draw()};
 var cf=$('clrfil');if(cf)cf.onclick=function(){secFilter='';grpFilter='';renderCtl();draw()};
 document.querySelectorAll('.modetog button').forEach(function(b){b.onclick=function(){MODE=b.dataset.m;sortKey=null;render()}});
}
function draw(){
 var p=P(), h=p.horizons[sel], rev=D.has_revision, M=m3(), tbl=$('tbl');
 var all=(h.rows||[]).filter(function(d){return (!secFilter||d.sec===secFilter)&&(!grpFilter||(d.grp||'').indexOf(grpFilter)>=0)});
 var rows=sortRows(all), shown=showAll?rows:rows.slice(0,TOP);
 var toggle=all.length>TOP?'<button class="lim" id="limtog">'+
  (showAll?'상위 '+TOP+'만 보기':'전체 '+all.length+'개 보기')+'</button>':'';
 var bK=M?'base3':'base', cK=M?'wow3':'wow', pK=M?'pwow3':'pwow';
 var bLab=M?'3개월전 컨센':'컨센 전주', bDate=M?p.ref3_date:p.base_date, pbDate=M?p.price_date_ref3:p.price_date_base;
 var cols=rev?
  [['mkt','시장','','mktc'],['sec','섹터','','l'],['name','종목','','l'],[bK,bLab,mmdd(bDate),''],
   ['curr','컨센 현재',mmdd(p.snapshot_date),''],[cK,'컨센 변화',mmdd(bDate)+'-'+mmdd(p.snapshot_date),''],
   [pK,'주가 변동',mmdd(pbDate)+'-'+mmdd(p.price_date_snap),'']]:
  [['mkt','시장','','mktc'],['sec','섹터','','l'],['name','종목','','l'],['curr','영업이익',mmdd(p.snapshot_date),'']];
 var out=toggle+'<table><tr><th>#</th>';
 cols.forEach(function(c){out+='<th class="sortable '+(c[3]||'')+'" data-k="'+c[0]+'">'+c[1]+
  (c[2]?'<br><span class="tiny">'+c[2]+'</span>':'')+arrow(c[0])+'</th>'});
 out+='</tr>';
 shown.forEach(function(d,i){out+='<tr><td>'+(i+1)+'</td>'+
  '<td class="mktc">'+mktBadge(d.mkt)+'</td>'+
  '<td class="l sec"><span class="secpick" data-s="'+d.sec+'">'+d.sec+'</span></td>'+
  '<td class="l">'+nameCell(d)+'</td>'+
  (rev?'<td>'+fmt(d[bK])+'</td><td>'+fmt(d.curr)+'</td><td>'+pct(d[cK])+'</td><td>'+pct(d[pK])+'</td>':'<td>'+fmt(d.curr)+'</td>')+'</tr>'});
 if(!shown.length)out+='<tr><td colspan="'+(cols.length+1)+'" style="text-align:center;padding:18px" class="tiny">해당 섹터 종목이 없습니다</td></tr>';
 tbl.innerHTML=out+'</table>';
 var lt=$('limtog');if(lt)lt.onclick=function(){showAll=!showAll;draw()};
 tbl.querySelectorAll('.secpick').forEach(function(el){el.onclick=function(){
  secFilter=(secFilter===el.dataset.s)?'':el.dataset.s;renderCtl();draw()}});   // 같은 섹터 재클릭=해제
 tbl.querySelectorAll('th.sortable').forEach(function(th){th.onclick=function(){
  var k=th.dataset.k;if(sortKey===k){sortDir=-sortDir}else{sortKey=k;sortDir=(k==='name'||k==='sec')?1:-1}draw()}});
}
function render(){renderNav();renderCtl();renderTop();draw()}
var tabs='';P().horizons.forEach(function(h,i){var lab=h.key==='q'?qq(h.period):h.label;
 tabs+='<button class="tab'+(i===0?' on':'')+'" data-i="'+i+'">'+lab+'</button>'});
$('tabs').innerHTML=tabs;
document.querySelectorAll('.tab').forEach(function(b){b.onclick=function(){
 sel=+b.dataset.i;document.querySelectorAll('.tab').forEach(function(x){x.classList.remove('on')});b.classList.add('on');draw()}});
render();
</script></body></html>"""


if __name__ == "__main__":
    build()
