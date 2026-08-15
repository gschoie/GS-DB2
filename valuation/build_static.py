# -*- coding: utf-8 -*-
"""valuation.json → 'valuation_report.html' + 'valuation_full.xlsx'.

화면은 셀사이드 피어 테이블 레이아웃이다 — 행=기업(산업 › 지역으로 묶음),
열=시가총액 + 지표(PER·EV/EBITDA·PBR·ROE·PSR)별 연도 묶음. 각 산업 블록 끝에
지역별·산업 전체 중앙값을 붙인다.

출력: telegram_research_dashboard/static/{valuation_report.html, valuation_full.xlsx}
"""

from __future__ import annotations

import json
import statistics
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

ROOT = Path(__file__).resolve().parent
STATIC = (ROOT.parent / "telegram_research_dashboard" / "static").resolve()
IN_JSON = ROOT / "data" / "valuation.json"
OUT_HTML = STATIC / "valuation_report.html"
OUT_XLSX = STATIC / "valuation_full.xlsx"

# 표시 순서. 참고 레이아웃(PER → EV/EBITDA → PBR → ROE)에 PSR을 덧붙였다.
METRICS = [
    ("per", "PER", "배"),
    ("ev_ebitda", "EV/EBITDA", "배"),
    ("pbr", "PBR", "배"),
    ("roe", "ROE", "%"),
    ("psr", "PSR", "배"),
]

NAVY = "1F4E79"
EST_FILL = "FFF6E5"
GROUP_FILL = "EDF1F6"
SUMMARY_FILL = "E8E0D2"


def year_axis(companies):
    """종목별 회계연도가 달라도 한 축에 세우도록 연도 합집합을 만든다.

    결산월이 다르면 같은 연도가 어떤 종목엔 확정(A)이고 다른 종목엔 컨센(E)이다
    — 3월 결산 일본 기업(고마쓰·히타치)의 FY2026은 이미 확정이지만 12월 결산
    한화에어로의 2026은 아직 컨센이다. 그래서 열의 A/E 표기는 '다수결'로 정하고,
    커버리지(그 연도 값을 가진 종목 수)를 함께 실어 보내 화면이 얇은 연도를
    기본 화면에서 걸러낼 수 있게 한다.
    """
    counts = {}
    for company in companies:
        for year in company.get("years", []):
            if year.get("unavailable"):
                continue          # 컨센 없는 빈 슬롯은 커버리지로 치지 않는다
            bucket = counts.setdefault(year["year"], {"A": 0, "E": 0})
            bucket[year["kind"]] += 1
    axis = []
    for year in sorted(counts):
        bucket = counts[year]
        kind = "A" if bucket["A"] >= bucket["E"] else "E"
        axis.append({"year": year, "kind": kind, "label": f"{year}{kind}",
                     "coverage": bucket["A"] + bucket["E"]})
    return axis


def value_of(company, year, metric):
    for entry in company.get("years", []):
        if entry["year"] == year:
            return entry.get(metric)
    return None


def median_of(companies, year, metric):
    """피어 요약은 평균이 아니라 중앙값을 쓴다.

    배수는 소수 종목(적자 직후 PER 수백배, 자본 얇은 방산주 PBR 수십배)이
    평균을 통째로 끌고 간다 — Boeing PBR 한 종목이 미국계 평균을 왜곡하는 식.
    중앙값은 그 왜곡 없이 '피어 허리'를 보여준다.
    """
    values = [v for v in (value_of(c, year, metric) for c in companies) if v is not None]
    return round(statistics.median(values), 2) if values else None


def industry_blocks(companies):
    """[(산업, [(지역, [기업…])…], 기업전체)] 순서를 원본 리스트 순서대로 유지."""
    blocks = []
    for company in companies:
        industry = company["industry"]
        block = next((b for b in blocks if b["industry"] == industry), None)
        if block is None:
            block = {"industry": industry, "regions": []}
            blocks.append(block)
        region = next((r for r in block["regions"] if r["region"] == company["region"]), None)
        if region is None:
            region = {"region": company["region"], "companies": []}
            block["regions"].append(region)
        region["companies"].append(company)
    for block in blocks:
        block["companies"] = [c for r in block["regions"] for c in r["companies"]]
    return blocks


# ─────────────────────────── 엑셀 ───────────────────────────

def build_xlsx(payload):
    companies = payload["companies"]
    axis = year_axis(companies)
    blocks = industry_blocks(companies)

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    head_font = Font(name="Malgun Gothic", size=10, bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor=NAVY)
    body_font = Font(name="Malgun Gothic", size=10)
    bold_font = Font(name="Malgun Gothic", size=10, bold=True)
    est_fill = PatternFill("solid", fgColor=EST_FILL)
    group_fill = PatternFill("solid", fgColor=GROUP_FILL)
    summary_fill = PatternFill("solid", fgColor=SUMMARY_FILL)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="BFBFBF")

    # ── 넓은 피어 테이블 시트 ──
    sheet = workbook.create_sheet("밸류에이션")
    sheet.cell(row=1, column=1, value="회사").font = head_font
    sheet.cell(row=1, column=1).fill = head_fill
    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    sheet.cell(row=1, column=2, value="시가총액").font = head_font
    sheet.cell(row=1, column=2).fill = head_fill
    sheet.cell(row=2, column=2, value="(백만달러)").font = head_font
    sheet.cell(row=2, column=2).fill = head_fill
    for cell in (sheet.cell(row=1, column=2), sheet.cell(row=2, column=2)):
        cell.alignment = center

    column = 3
    metric_spans = []
    for key, label, unit in METRICS:
        start = column
        for entry in axis:
            head = sheet.cell(row=2, column=column, value=entry["label"])
            head.font, head.fill, head.alignment = head_font, head_fill, center
            if entry["kind"] == "E":
                head.border = Border(bottom=thin)
            column += 1
        sheet.merge_cells(start_row=1, start_column=start, end_row=1, end_column=column - 1)
        title = sheet.cell(row=1, column=start, value=f"{label} ({unit})")
        title.font, title.fill, title.alignment = head_font, head_fill, center
        metric_spans.append((key, start))

    row_index = 3
    for block in blocks:
        cell = sheet.cell(row=row_index, column=1, value=block["industry"])
        cell.font = bold_font
        for c in range(1, column):
            sheet.cell(row=row_index, column=c).fill = group_fill
        row_index += 1

        for region in block["regions"]:
            cell = sheet.cell(row=row_index, column=1, value=f"  {region['region']}")
            cell.font = bold_font
            row_index += 1
            for company in region["companies"]:
                sheet.cell(row=row_index, column=1, value=f"    {company['name']}").font = body_font
                cap = sheet.cell(row=row_index, column=2, value=company.get("market_cap_musd"))
                cap.font, cap.number_format = body_font, "#,##0"
                for key, start in metric_spans:
                    for offset, entry in enumerate(axis):
                        target = sheet.cell(row=row_index, column=start + offset,
                                            value=value_of(company, entry["year"], key))
                        target.font = body_font
                        target.number_format = "0.0" if key == "roe" else "0.00"
                        if entry["kind"] == "E":
                            target.fill = est_fill
                row_index += 1

        # 산업 블록 끝: 지역별 + 산업 전체 중앙값
        for region in block["regions"]:
            _write_summary(sheet, row_index, column, metric_spans, axis,
                           f"  {region['region']} 중앙값", region["companies"],
                           bold_font, summary_fill)
            row_index += 1
        _write_summary(sheet, row_index, column, metric_spans, axis,
                       f"{block['industry']} 전체 중앙값", block["companies"],
                       bold_font, summary_fill)
        row_index += 2

    sheet.freeze_panes = "C3"
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 12
    for index in range(3, column):
        sheet.column_dimensions[get_column_letter(index)].width = 9

    # ── 지표별 시트(차트·정렬용) ──
    for key, label, unit in METRICS:
        detail = workbook.create_sheet(f"{label.replace('/', '_')}({unit})")
        headers = ["기업명", "티커", "산업", "지역", "시총(백만$)"] + [a["label"] for a in axis]
        for index, name in enumerate(headers, 1):
            cell = detail.cell(row=1, column=index, value=name)
            cell.font, cell.fill, cell.alignment = head_font, head_fill, center
        for offset, company in enumerate(companies, 2):
            values = [company["name"], company["ticker"], company["industry"], company["region"],
                      company.get("market_cap_musd")]
            values += [value_of(company, a["year"], key) for a in axis]
            for index, value in enumerate(values, 1):
                cell = detail.cell(row=offset, column=index, value=value)
                cell.font = body_font
                if index >= 6:
                    cell.number_format = "0.0" if key == "roe" else "0.00"
        detail.freeze_panes = "F2"
        for index, width in enumerate([18, 11, 10, 10, 12] + [10] * len(axis), 1):
            detail.column_dimensions[get_column_letter(index)].width = width

    # ── 원자료 ──
    raw = workbook.create_sheet("원자료")
    raw_columns = ["기업명", "티커", "산업", "지역", "연도", "구분", "주가", "EPS", "BPS",
                   "매출액", "순이익", "자본총계", "PER", "EV/EBITDA", "PBR", "ROE(%)", "PSR",
                   "주가통화", "재무통화"]
    for index, name in enumerate(raw_columns, 1):
        cell = raw.cell(row=1, column=index, value=name)
        cell.font, cell.fill, cell.alignment = head_font, head_fill, center
    row_index = 2
    for company in companies:
        for entry in company.get("years", []):
            values = [
                company["name"], company["ticker"], company["industry"], company["region"],
                entry["label"], "확정" if entry["kind"] == "A" else "컨센",
                entry.get("price"), entry.get("eps"), entry.get("bps"),
                entry.get("revenue"), entry.get("net_income"), entry.get("equity"),
                entry.get("per"), entry.get("ev_ebitda"), entry.get("pbr"),
                entry.get("roe"), entry.get("psr"),
                company.get("currency"), company.get("fin_currency"),
            ]
            for index, value in enumerate(values, 1):
                cell = raw.cell(row=row_index, column=index, value=value)
                cell.font = body_font
                if index in (10, 11, 12):
                    cell.number_format = "#,##0"
            if entry["kind"] == "E":
                for index in range(1, len(raw_columns) + 1):
                    raw.cell(row=row_index, column=index).fill = est_fill
            row_index += 1
    raw.freeze_panes = "A2"
    for index, width in enumerate([18, 11, 10, 10, 9, 8] + [13] * 13, 1):
        raw.column_dimensions[get_column_letter(index)].width = width

    # ── 산출기준 ──
    notes = workbook.create_sheet("산출기준")
    for index, line in enumerate([
        ["항목", "정의"],
        ["시가총액", "현재 시총을 백만달러로 환산(야후 환율)"],
        ["PER (확정)", "회계연도 기말 시총 ÷ 지배주주순이익"],
        ["PBR (확정)", "회계연도 기말 시총 ÷ 자본총계(기말)"],
        ["PSR (확정)", "회계연도 기말 시총 ÷ 매출액"],
        ["ROE (확정)", "순이익 ÷ 자본총계(기초·기말 평균, 기초 없으면 기말)"],
        ["EV/EBITDA (확정)", "(시총 + 총차입 − 현금) ÷ EBITDA. EBITDA 행이 없으면 영업이익 + 감가상각"],
        ["PER (컨센)", "현재주가 ÷ 야후 컨센서스 EPS"],
        ["PSR (컨센)", "현재시총 ÷ 야후 컨센서스 매출액"],
        ["PBR (컨센)", "현재시총 ÷ 예상자본 — ※ 컨센 원본 아님(이익잉여금 롤포워드 자체 산출)"],
        ["ROE (컨센)", "컨센순이익 ÷ 평균예상자본 — ※ 컨센 원본 아님(자체 산출)"],
        ["EV/EBITDA (컨센)", "야후에 EBITDA 컨센이 없어 공란"],
        ["예상자본", "직전 확정자본 + 컨센순이익 × (1 − 배당성향)"],
        ["2년후", "야후 파이낸스가 2년후 컨센서스를 제공하지 않아 공란"],
        ["요약행", "평균이 아니라 중앙값 — 극단 배수 종목이 평균을 왜곡하기 때문"],
        ["통화", "주가통화 ≠ 재무통화면 환율로 환산(예: 두산밥캣 호가 KRW / 재무 USD)"],
        ["적자·자본잠식", "분모가 0 이하면 배수를 만들지 않고 공란. 단 ROE 음수는 그대로 표기"],
        ["출처", "Yahoo Finance (yfinance)"],
        ["기준시각", payload.get("updated_at", "")],
    ], 1):
        for column, value in enumerate(line, 1):
            cell = notes.cell(row=index, column=column, value=value)
            cell.font = head_font if index == 1 else body_font
            if index == 1:
                cell.fill = head_fill
    notes.column_dimensions["A"].width = 20
    notes.column_dimensions["B"].width = 82

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUT_XLSX)
    print(f"📥 엑셀: {OUT_XLSX} ({OUT_XLSX.stat().st_size / 1024:.0f} KB)")


def _write_summary(sheet, row_index, end_column, metric_spans, axis, label, companies,
                   font, fill):
    cell = sheet.cell(row=row_index, column=1, value=label)
    cell.font = font
    for column in range(1, end_column):
        sheet.cell(row=row_index, column=column).fill = fill
    for key, start in metric_spans:
        for offset, entry in enumerate(axis):
            target = sheet.cell(row=row_index, column=start + offset,
                                value=median_of(companies, entry["year"], key))
            target.font = font
            target.number_format = "0.0" if key == "roe" else "0.00"
            target.fill = fill


# ─────────────────────────── 화면 ───────────────────────────

HTML_TEMPLATE = """<!doctype html>
<html lang="ko"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>피어그룹 Valuation</title>
<style>
:root{--bg:#f6f7f9;--card:#fff;--ink:#1a1d21;--muted:#6b7280;--line:#e5e7eb;
--accent:#185fa5;--est:#fff6e5;--grp:#eef2f7;--sum:#efe8dc;--bad:#d03b3b}
@media(prefers-color-scheme:dark){:root{--bg:#141516;--card:#1d1e1f;--ink:#f2f2f0;
--muted:#9a9a92;--line:#2f3032;--est:#2a2418;--grp:#232629;--sum:#2c2822}}
*{box-sizing:border-box}body{margin:0;background:var(--bg);color:var(--ink);
font-family:-apple-system,"Segoe UI","Malgun Gothic",sans-serif;line-height:1.45}
.wrap{max-width:1600px;margin:0 auto;padding:20px 16px 60px}
h1{font-size:19px;font-weight:600;margin:0 0 2px}
.sub{color:var(--muted);font-size:12.5px}
.bar{display:flex;flex-wrap:wrap;gap:8px;align-items:center;margin:14px 0 12px}
.btn{display:inline-flex;align-items:center;gap:6px;background:var(--accent);color:#fff;
text-decoration:none;font-size:13px;font-weight:600;padding:9px 15px;border-radius:9px;
border:0;cursor:pointer;font-family:inherit}
.btn.ghost{background:var(--card);color:var(--ink);border:1px solid var(--line)}
.btn:disabled{opacity:.55;cursor:default}
input[type=search]{flex:1;min-width:140px;background:var(--card);border:1px solid var(--line);
border-radius:9px;padding:9px 12px;font-size:13px;color:var(--ink)}
.chips{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:10px;align-items:center}
.chip{background:var(--card);border:1px solid var(--line);border-radius:16px;padding:5px 13px;
font-size:12px;color:var(--muted);cursor:pointer;font-family:inherit}
.chip.on{background:var(--ink);border-color:var(--ink);color:var(--bg)}
.chip.mt.on{background:var(--accent);border-color:var(--accent);color:#fff}
.chips .gap{width:10px}
#status{font-size:12.5px;color:var(--muted);min-height:18px;margin-bottom:8px}
.tablewrap{background:var(--card);border:1px solid var(--line);border-radius:10px;overflow:auto}
table{border-collapse:separate;border-spacing:0;font-size:12.5px;white-space:nowrap}
th,td{padding:5px 8px;text-align:right;border-bottom:1px solid var(--line)}
thead th{background:var(--card);position:sticky;top:0;z-index:2;font-weight:600;
font-size:11.5px;color:var(--muted)}
thead tr:first-child th{top:0}thead tr:last-child th{top:26px}
th.name,td.name{text-align:left;position:sticky;left:0;background:var(--card);z-index:1;
min-width:190px}
thead th.name{z-index:3}
th.mg{border-left:1px solid var(--line)}td.mg{border-left:1px solid var(--line)}
td.est,th.est{background:var(--est)}
tr.grp td{background:var(--grp);font-weight:700;font-size:12px}
tr.grp td.name{background:var(--grp)}
tr.reg td{background:transparent;font-weight:600;color:var(--muted);font-size:11.5px}
tr.sum td{background:var(--sum);font-weight:600}
tr.sum td.name{background:var(--sum)}
td.name .tk{color:var(--muted);font-size:11px;margin-left:6px}
tr.co:hover td{background:rgba(24,95,165,.07)}
tr.co:hover td.name{background:rgba(24,95,165,.07)}
.na{color:var(--muted)}
.mark{color:var(--accent);font-weight:700}
.legend{margin-top:12px;font-size:11.5px;color:var(--muted);line-height:1.75}
.legend b{color:var(--ink)}
.warn{margin-top:8px;font-size:11.5px;color:var(--bad)}
</style></head><body><div class="wrap">
<h1>피어그룹 Valuation — PER · EV/EBITDA · PBR · ROE · PSR</h1>
<div class="sub">야후 파이낸스 기준 · 확정 __HY__개 회계연도 + 컨센(올해·내년) · 기준 __UPDATED__</div>

<div class="bar">
  <a class="btn" href="valuation_full.xlsx" download>📥 엑셀 다운로드</a>
  <button class="btn" id="run" type="button">▶ 수집 실행</button>
  <button class="btn ghost" id="csv" type="button">⤓ 현재 화면 CSV</button>
  <input type="search" id="q" placeholder="기업명 · 티커 검색">
</div>
<div id="status"></div>

<div class="chips" id="chips"></div>
<div class="tablewrap"><table><thead id="head"></thead><tbody id="body"></tbody></table></div>

<div class="legend">
  <b>확정(A)</b> = 회계연도 기말 주가·재무제표 기준 · <b>컨센(E)</b> = 음영 열, 현재 주가 기준 ·
  <b>시가총액</b> = 백만달러 환산<br>
  PER(E)·PSR(E)는 야후 컨센서스 EPS·매출 원본. <b>PBR(E)·ROE(E)는 컨센 항목이 없어</b>
  «직전 확정자본 + 컨센순이익 × (1 − 배당성향)» 롤포워드로 <b>자체 산출</b>했다(<span class="mark">*</span> 표기).
  EV/EBITDA는 야후에 EBITDA 컨센이 없어 컨센 열이 비어 있고, 2년후 열은 컨센 자체가 없어 비어 있다.<br>
  <b>요약행은 평균이 아니라 중앙값</b>이다 — 적자 직후 PER 수백배·자본 얇은 종목 PBR 수십배 같은
  극단값 하나가 평균을 통째로 끌고 가기 때문. 적자·자본잠식으로 분모가 0 이하인 배수는 공란(ROE 음수는 표기).
</div>
<div class="warn" id="warn"></div>
</div>
<script>
const DATA = __PAYLOAD__;
const METRICS = __METRICS__;
const AXIS = __AXIS__;
const BLOCKS = __BLOCKS__;
/* 갱신 버튼 → GitHub Actions 디스패치 GAS 웹앱 (gas/dispatch_proxy.gs).
   workflow 키 'valuation' 은 그 파일의 WF 매핑과 1:1이어야 한다. */
const DISPATCH_ENDPOINT='https://script.google.com/macros/s/AKfycbx3RjIjtlO2Z6fIYo2T3LhJrFg9Wp2hS7dMS3Is52-JVF1hizoCWewbQ1uM_v5sdhR2jw/exec';

const state = {industry:'전체', q:'', span:'ref', metrics:METRICS.map(m=>m[0])};
const industries = ['전체', ...new Set(DATA.companies.map(c=>c.industry))];

/* 지표 5종을 다 켜면 표가 넓다 — 보고 싶은 것만 켜서 좁게 볼 수 있게 한다.
   순서는 항상 METRICS 순서를 따른다(켠 순서대로 뒤섞이지 않게). */
function activeMetrics(){
  return METRICS.filter(m=>state.metrics.includes(m[0]));
}

/* 기본은 참고 레이아웃과 같은 4열(확정 2 + 컨센 2).
   결산월이 다른 소수 종목 때문에 생긴 얇은 연도(예: 3월 결산만 값이 있는 해)는
   기본 화면에서 뺀다 — 대부분 칸이 빈 열이 끼면 표가 읽히지 않는다.
   '전체 연도'를 켜면 그 열까지 전부 보여준다. */
function axisFor(){
  if(state.span === 'all') return AXIS;
  const need = Math.max(2, DATA.companies.length * 0.5);
  const solid = AXIS.filter(x=>x.coverage >= need);
  return [...solid.filter(x=>x.kind==='A').slice(-2),
          ...solid.filter(x=>x.kind==='E').slice(0, 2)];
}
const fmt = (v,m) => v==null ? '' : (m==='roe' ? v.toFixed(1) : v.toFixed(2));

function visible(){
  const q = state.q.trim().toLowerCase();
  return DATA.companies.filter(c =>
    (state.industry==='전체' || c.industry===state.industry) &&
    (!q || c.name.toLowerCase().includes(q) || c.ticker.toLowerCase().includes(q)));
}

function median(list, year, metric){
  const vals = list.map(c=>{
    const y=(c.years||[]).find(y=>y.year===year); return y ? y[metric] : null;
  }).filter(v=>v!=null).sort((a,b)=>a-b);
  if(!vals.length) return null;
  const mid = Math.floor(vals.length/2);
  return vals.length%2 ? vals[mid] : (vals[mid-1]+vals[mid])/2;
}

function cells(getter, axis, cls){
  let html='';
  for(const [mi,m] of activeMetrics().entries())
    for(const [ai,a] of axis.entries())
      html += `<td class="${ai===0&&mi>0?'mg':''} ${a.kind==='E'?'est':''} ${cls||''}">`+
              `${getter(m[0], a)}</td>`;
  return html;
}

function render(){
  const axis = axisFor(), metrics = activeMetrics();
  document.getElementById('chips').innerHTML =
    industries.map(i=>`<button class="chip ${i===state.industry?'on':''}" data-i="${i}">${i}</button>`).join('')
    + '<span class="gap"></span>'
    + `<button class="chip ${state.span==='ref'?'on':''}" data-s="ref">최근 4열</button>`
    + `<button class="chip ${state.span==='all'?'on':''}" data-s="all">전체 연도</button>`
    + '<span class="gap"></span>'
    + METRICS.map(m=>`<button class="chip mt ${state.metrics.includes(m[0])?'on':''}" data-m="${m[0]}">${m[1]}</button>`).join('')
    + `<button class="chip" data-m="__all">전체 지표</button>`;

  let h1 = '<tr><th class="name" rowspan="2">회사</th><th rowspan="2">시가총액<br><span style="font-weight:400">(백만$)</span></th>';
  let h2 = '<tr>';
  for(const [mi,m] of metrics.entries()){
    h1 += `<th class="${mi>0?'mg':''}" colspan="${axis.length}">${m[1]}${m[2]==='%'?' (%)':''}</th>`;
    for(const [ai,a] of axis.entries())
      h2 += `<th class="${ai===0&&mi>0?'mg':''} ${a.kind==='E'?'est':''}">${a.label}</th>`;
  }
  document.getElementById('head').innerHTML = h1+'</tr>'+h2+'</tr>';

  const rows = visible();
  const span = 2 + metrics.length*axis.length;
  let html = '';
  for(const block of BLOCKS){
    const inBlock = rows.filter(c=>c.industry===block.industry);
    if(!inBlock.length) continue;
    html += `<tr class="grp"><td class="name">${block.industry}</td><td colspan="${span-1}"></td></tr>`;
    for(const region of block.regions){
      const list = inBlock.filter(c=>c.region===region.region);
      if(!list.length) continue;
      html += `<tr class="reg"><td class="name">${region.region}</td><td colspan="${span-1}"></td></tr>`;
      for(const c of list){
        html += `<tr class="co"><td class="name">${c.name}<span class="tk">${c.ticker}</span></td>`+
                `<td>${c.market_cap_musd!=null?c.market_cap_musd.toLocaleString():''}</td>`+
                cells((m,a)=>{
                  const y=(c.years||[]).find(y=>y.year===a.year);
                  const v=y?y[m]:null;
                  const der = y && (y.derived||[]).includes(m) && v!=null;
                  return v==null ? '<span class="na"></span>'
                                 : fmt(v,m)+(der?'<span class="mark">*</span>':'');
                }, axis) + '</tr>';
      }
    }
    for(const region of block.regions){
      const list = inBlock.filter(c=>c.region===region.region);
      if(list.length<2) continue;
      html += `<tr class="sum"><td class="name">${region.region} 중앙값</td><td></td>`+
              cells((m,a)=>fmt(median(list,a.year,m),m), axis) + '</tr>';
    }
    html += `<tr class="sum"><td class="name">${block.industry} 전체 중앙값</td><td></td>`+
            cells((m,a)=>fmt(median(inBlock,a.year,m),m), axis) + '</tr>';
  }
  document.getElementById('body').innerHTML = html ||
    `<tr><td class="name">조건에 맞는 종목이 없습니다.</td><td colspan="${span-1}"></td></tr>`;

  const bad = DATA.companies.filter(c=>(c.warnings||[]).some(w=>w.includes('못')));
  document.getElementById('warn').innerHTML = bad.length
    ? '⚠️ ' + bad.map(c=>`${c.name}: ${c.warnings[0]}`).join(' / ') : '';
}

document.getElementById('chips').addEventListener('click', e=>{
  const b = e.target.closest('.chip'); if(!b) return;
  if(b.dataset.i) state.industry = b.dataset.i;
  if(b.dataset.s) state.span = b.dataset.s;
  if(b.dataset.m === '__all') state.metrics = METRICS.map(m=>m[0]);
  else if(b.dataset.m){
    const key = b.dataset.m, on = state.metrics.includes(key);
    // 마지막 한 개는 끄지 않는다 — 지표가 0개면 볼 게 없는 표가 된다.
    if(!(on && state.metrics.length === 1))
      state.metrics = on ? state.metrics.filter(k=>k!==key) : [...state.metrics, key];
  }
  render();
});
document.getElementById('q').addEventListener('input', e=>{ state.q = e.target.value; render(); });

/* 수집 실행 — GAS 프록시가 valuation.yml 을 workflow_dispatch 로 돌린다.
   결과 반영까지 수집(1~2분) + Pages 배포(2~3분)가 필요해 안내를 함께 띄운다. */
document.getElementById('run').addEventListener('click', async ()=>{
  const btn = document.getElementById('run'), status = document.getElementById('status');
  btn.disabled = true; status.textContent = '요청 중…';
  try{
    const r = await fetch(DISPATCH_ENDPOINT, {method:'POST', body:JSON.stringify({workflow:'valuation'})});
    let denied = null;
    try{ const d = await r.json(); if(d && d.ok===false) denied = d; }catch{}
    status.textContent = denied
      ? `⚠ 거절 ${denied.code||'?'}${denied.wf?` (${denied.wf})`:''} — ${denied.error||'GAS 프록시 매핑 확인 필요'}`
      : '✅ 수집을 시작했습니다. 1~2분 뒤 수집이 끝나고, 배포까지 3~5분 뒤 새로고침하면 반영됩니다.';
  }catch(e){ status.textContent = '실패: ' + e.message; }
  finally{ btn.disabled = false; }
});

document.getElementById('csv').addEventListener('click', ()=>{
  const axis = axisFor();
  const cell = v => `"${String(v??'').replace(/"/g,'""')}"`;
  const head = ['기업명','티커','산업','지역','시총(백만$)'];
  for(const m of activeMetrics()) for(const a of axis) head.push(`${m[1]} ${a.label}`);
  const rows = visible().map(c=>{
    const row = [c.name, c.ticker, c.industry, c.region, c.market_cap_musd ?? ''];
    for(const m of activeMetrics()) for(const a of axis){
      const y=(c.years||[]).find(y=>y.year===a.year); row.push(y ? (y[m[0]] ?? '') : '');
    }
    return row;
  });
  const csv = '\\ufeff' + [head, ...rows].map(r=>r.map(cell).join(',')).join('\\r\\n');
  const url = URL.createObjectURL(new Blob([csv], {type:'text/csv;charset=utf-8'}));
  const a = document.createElement('a');
  a.href = url; a.download = `피어그룹밸류에이션_${DATA.updated_at.slice(0,10)}.csv`;
  document.body.appendChild(a); a.click(); a.remove();
  setTimeout(()=>URL.revokeObjectURL(url), 1000);
});

render();
</script></body></html>
"""


def build_html(payload):
    companies = payload["companies"]
    axis = year_axis(companies)
    blocks = [{"industry": b["industry"],
               "regions": [{"region": r["region"]} for r in b["regions"]]}
              for b in industry_blocks(companies)]
    data = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")
    html = (HTML_TEMPLATE
            .replace("__PAYLOAD__", data)
            .replace("__METRICS__", json.dumps([[k, l, u] for k, l, u in METRICS], ensure_ascii=False))
            .replace("__AXIS__", json.dumps(axis, ensure_ascii=False))
            .replace("__BLOCKS__", json.dumps(blocks, ensure_ascii=False))
            .replace("__HY__", str(payload.get("history_years", 4)))
            .replace("__UPDATED__", payload.get("updated_at", "")[:16].replace("T", " ")))
    OUT_HTML.parent.mkdir(parents=True, exist_ok=True)
    OUT_HTML.write_text(html, encoding="utf-8")
    print(f"🖥️  화면: {OUT_HTML} ({OUT_HTML.stat().st_size / 1024:.0f} KB)")


def main():
    if not IN_JSON.exists():
        sys.exit(f"수집 데이터가 없습니다: {IN_JSON} — 먼저 fetch_valuation.py 를 실행하세요.")
    payload = json.loads(IN_JSON.read_text(encoding="utf-8"))
    if not payload.get("companies"):
        sys.exit("수집 결과가 비어 있어 산출물을 만들지 않습니다.")
    build_xlsx(payload)
    build_html(payload)


if __name__ == "__main__":
    main()
